from openpyxl import Workbook

wb = Workbook()
ws = wb.active

with open('lat.txt', 'r') as f:
    lat = f.readlines()
with open('long.txt', 'r') as f2:
    lg = f2.readlines()


ws['A1'] = 'latitude'
ws['B1'] = 'longitude'
count = 1
count2 = 1
for i in lat:
    count += 1
    ws[f'A{count}'] = i

for j in lg:
    count2 += 1
    ws[f'B{count2}'] = j

wb.save('sample.xlsx')
